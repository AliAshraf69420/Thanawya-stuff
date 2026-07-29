from datetime import datetime
from openpyxl import load_workbook

# ===========================
# Configuration
# ===========================
FILE = "thanawya_grades.xlsx"
OUTPUT_FILE = "output.md"

SEARCH_COLUMN = "arabic_name"

# ===========================
# Load workbook
# ===========================
wb = load_workbook(FILE, read_only=True)
ws = wb.active

rows = ws.iter_rows(values_only=True)
headers = list(next(rows))

if SEARCH_COLUMN not in headers:
    raise ValueError(f"Column '{SEARCH_COLUMN}' not found.")

search_idx = headers.index(SEARCH_COLUMN)

# ===========================
# Search
# ===========================
search_name = input("Enter student name: ").strip()

matches = []

for row in rows:
    value = row[search_idx]

    if value is None:
        continue

    if search_name.lower() in str(value).strip().lower():
        matches.append(row)

# ===========================
# Write results
# ===========================
timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

with open(OUTPUT_FILE, "a", encoding="utf-8") as f:
    # Add spacing if the file already has content
    f.seek(0, 2)
    if f.tell() != 0:
        f.write("\n\n---\n\n")

    f.write(f"# Search Query: `{search_name}`\n\n")
    f.write(f"**Timestamp:** {timestamp}\n\n")
    f.write(f"**Matches Found:** {len(matches)}\n\n")

    if not matches:
        f.write("No matching students were found.\n")
    else:
        for i, row in enumerate(matches, start=1):
            f.write(f"## Match {i}\n\n")
            f.write("| Field | Value |\n")
            f.write("|:------|:------|\n")

            for header, cell in zip(headers, row):
                value = "" if cell is None else str(cell).replace("\n", " ")
                # Escape pipes so Markdown tables don't break
                value = value.replace("|", "\\|")
                f.write(f"| {header} | {value} |\n")

            f.write("\n")

print(f"\nFound {len(matches)} match(es).")
print(f"Results appended to '{OUTPUT_FILE}'.")