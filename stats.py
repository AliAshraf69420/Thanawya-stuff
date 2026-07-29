from collections import Counter
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt

# ===========================
# Configuration
# ===========================
FILE = "thanawya_grades.xlsx"
GRADE_COLUMN = "total_degree"

# ===========================
# Read grades
# ===========================
wb = load_workbook(FILE, read_only=True)
ws = wb.active

rows = ws.iter_rows(values_only=True)
headers = list(next(rows))

if GRADE_COLUMN not in headers:
    raise ValueError(f"Column '{GRADE_COLUMN}' not found.")

grade_idx = headers.index(GRADE_COLUMN)

grades = []

for row in rows:
    value = row[grade_idx]

    if value is None or value == "":
        continue

    try:
        grades.append(float(value))
    except (TypeError, ValueError):
        pass

if not grades:
    raise ValueError("No valid grades found.")

grades = np.array(grades)

# ===========================
# Statistics
# ===========================
mean = np.mean(grades)
median = np.median(grades)

q1 = np.percentile(grades, 25)
q2 = np.percentile(grades, 50)
q3 = np.percentile(grades, 75)

minimum = np.min(grades)
maximum = np.max(grades)

counter = Counter(grades)
highest_frequency = max(counter.values())
modes = sorted([g for g, c in counter.items() if c == highest_frequency])

print("=" * 50)
print("Grade Statistics")
print("=" * 50)
print(f"Students : {len(grades)}")
print(f"Mean     : {mean:.2f}")
print(f"Median   : {median:.2f}")

print("\nQuartiles")
print(f"Q1 (25%) : {q1:.2f}")
print(f"Q2 (50%) : {q2:.2f}")
print(f"Q3 (75%) : {q3:.2f}")

print(f"\nMinimum  : {minimum:.2f}")
print(f"Maximum  : {maximum:.2f}")

if highest_frequency == 1:
    print("\nMode     : No unique mode")
else:
    print(f"\nMode(s)  : {modes}")
    print(f"Frequency: {highest_frequency}")

# ===========================
# Histogram
# ===========================
plt.figure(figsize=(10, 6))

plt.hist(grades, bins=20)

plt.axvline(mean, linestyle="--", linewidth=2, label=f"Mean ({mean:.2f})")
plt.axvline(median, linestyle="-", linewidth=2, label=f"Median ({median:.2f})")

plt.title("Distribution of Student Grades")
plt.xlabel("Total Degree")
plt.ylabel("Number of Students")

plt.legend()
plt.tight_layout()

plt.savefig("grade_distribution.png", dpi=300)

plt.show()

print("\nHistogram saved as grade_distribution.png")