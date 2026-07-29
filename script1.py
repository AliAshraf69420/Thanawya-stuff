from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

FILE = "thanawya_grades.xlsx"

wb = load_workbook(FILE, read_only=True)
ws = wb.active  # or wb["SheetName"]

headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

print("=" * 80)
print("Columns in first row")
print("=" * 80)

for i, header in enumerate(headers, start=1):
    col_letter = get_column_letter(i)
    print(f"{i:3}. {col_letter:>3} | {repr(header)}")

print("=" * 80)