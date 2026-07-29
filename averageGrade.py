from openpyxl import load_workbook

FILE = "thanawya_grades.xlsx"

wb = load_workbook(FILE, read_only=True)
ws = wb.active

rows = ws.iter_rows(values_only=True)
headers = list(next(rows))

GRADE_COLUMN = "total_degree"

if GRADE_COLUMN not in headers:
    raise ValueError(f"Column '{GRADE_COLUMN}' not found.")

grade_idx = headers.index(GRADE_COLUMN)

total = 0.0
count = 0

for row in rows:
    grade = row[grade_idx]

    if grade is None or grade == "":
        continue

    try:
        grade = float(grade)
    except (TypeError, ValueError):
        continue

    total += grade
    count += 1

if count == 0:
    print("No valid grades found.")
else:
    average = total / count

    print("=" * 40)
    print(f"Students processed : {count}")
    print(f"Total grades       : {total:.2f}")
    print(f"Class average      : {average:.2f}")
    print("=" * 40)